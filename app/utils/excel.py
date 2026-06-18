import io
import csv
import openpyxl
import xlrd
from typing import List, Dict, Optional, Tuple

def extract_row_image(ws, row_idx_0_based: int) -> Optional[bytes]:
    """
    Search ws._images for an image anchored to row_idx_0_based.
    Supports both OneCellAnchor/TwoCellAnchor coordinate systems and string anchors (e.g. 'A2').
    """
    if not hasattr(ws, "_images") or not ws._images:
        return None
        
    for img in ws._images:
        # 1. Check cell coordinate string anchors (e.g. 'E2')
        if isinstance(img.anchor, str):
            try:
                from openpyxl.utils import coordinate_to_tuple
                col, row = coordinate_to_tuple(img.anchor)
                if row - 1 == row_idx_0_based:
                    # In openpyxl, row is 1-based, convert to 0-based
                    return img._data()
            except Exception:
                pass
        # 2. Check anchor object markers (OneCellAnchor / TwoCellAnchor)
        elif img.anchor and hasattr(img.anchor, "_from") and img.anchor._from:
            if img.anchor._from.row == row_idx_0_based:
                try:
                    return img._data()
                except Exception:
                    pass
    return None

def parse_student_file(file_content: bytes, filename: str) -> List[Dict]:
    """
    Parse uploaded files (xlsx, xlsm, xls, csv) and return student dict list:
    [
      {
        "scholar_number": "...",
        "name": "...",
        "email": "...",
        "contact_number": "...",
        "image_bytes": bytes or None
      }
    ]
    """
    fn_lower = filename.lower()
    students = []

    # 1. XLSX / XLSM (openpyxl)
    if fn_lower.endswith(".xlsx") or fn_lower.endswith(".xlsm"):
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active
        
        # Identify headers
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val:
                val_str = str(val).strip().lower()
                if "scholar" in val_str or "roll" in val_str or "admission" in val_str:
                    headers["scholar_number"] = col_idx
                elif "name" in val_str:
                    headers["name"] = col_idx
                elif "email" in val_str or "e-mail" in val_str:
                    headers["email"] = col_idx
                elif "contact" in val_str or "phone" in val_str or "mobile" in val_str or "number" in val_str:
                    # Make sure not to match scholar_number here
                    if "scholar" not in val_str:
                        headers["contact_number"] = col_idx

        # Fallback to column index order if headers are missing
        scholar_col = headers.get("scholar_number", 1)
        name_col = headers.get("name", 2)
        email_col = headers.get("email", 3)
        contact_col = headers.get("contact_number", 4)

        for row_idx in range(2, ws.max_row + 1):
            scholar_val = ws.cell(row=row_idx, column=scholar_col).value
            name_val = ws.cell(row=row_idx, column=name_col).value
            email_val = ws.cell(row=row_idx, column=email_col).value
            contact_val = ws.cell(row=row_idx, column=contact_col).value

            # Skip rows where core info is missing
            if scholar_val is None or name_val is None:
                continue

            scholar_number = str(scholar_val).strip()
            # Clean floating numbers in scholar number (e.g. 101.0 -> 101)
            if scholar_number.endswith(".0"):
                scholar_number = scholar_number[:-2]

            name = str(name_val).strip()
            email = str(email_val).strip() if email_val is not None else ""
            
            contact_number = str(contact_val).strip() if contact_val is not None else ""
            if contact_number.endswith(".0"):
                contact_number = contact_number[:-2]

            # Try to extract image anchored to this row (row_idx - 1 because ws._images anchor row is 0-based)
            image_bytes = extract_row_image(ws, row_idx - 1)

            students.append({
                "scholar_number": scholar_number,
                "name": name,
                "email": email,
                "contact_number": contact_number,
                "image_bytes": image_bytes
            })

    # 2. XLS (xlrd)
    elif fn_lower.endswith(".xls"):
        book = xlrd.open_workbook(file_contents=file_content)
        sheet = book.sheet_by_index(0)

        # Identify headers
        headers = {}
        for col_idx in range(sheet.ncols):
            val = sheet.cell_value(rowx=0, colx=col_idx)
            if val:
                val_str = str(val).strip().lower()
                if "scholar" in val_str or "roll" in val_str or "admission" in val_str:
                    headers["scholar_number"] = col_idx
                elif "name" in val_str:
                    headers["name"] = col_idx
                elif "email" in val_str or "e-mail" in val_str:
                    headers["email"] = col_idx
                elif "contact" in val_str or "phone" in val_str or "mobile" in val_str or "number" in val_str:
                    if "scholar" not in val_str:
                        headers["contact_number"] = col_idx

        scholar_col = headers.get("scholar_number", 0)
        name_col = headers.get("name", 1)
        email_col = headers.get("email", 2)
        contact_col = headers.get("contact_number", 3)

        for row_idx in range(1, sheet.nrows):
            scholar_val = sheet.cell_value(rowx=row_idx, colx=scholar_col)
            name_val = sheet.cell_value(rowx=row_idx, colx=name_col)
            email_val = sheet.cell_value(rowx=row_idx, colx=email_col)
            contact_val = sheet.cell_value(rowx=row_idx, colx=contact_col)

            if scholar_val == "" or name_val == "":
                continue

            scholar_number = str(scholar_val).strip()
            if scholar_number.endswith(".0"):
                scholar_number = scholar_number[:-2]

            name = str(name_val).strip()
            email = str(email_val).strip()
            
            contact_number = str(contact_val).strip()
            if contact_number.endswith(".0"):
                contact_number = contact_number[:-2]

            students.append({
                "scholar_number": scholar_number,
                "name": name,
                "email": email,
                "contact_number": contact_number,
                "image_bytes": None  # xlrd cannot parse embedded images
            })

    # 3. CSV (csv)
    elif fn_lower.endswith(".csv"):
        try:
            csv_text = file_content.decode("utf-8")
        except UnicodeDecodeError:
            # Fallback to latin-1
            csv_text = file_content.decode("latin-1")

        reader = csv.reader(io.StringIO(csv_text))
        rows = list(reader)
        if not rows:
            return []

        headers = {}
        first_row = rows[0]
        for col_idx, val in enumerate(first_row):
            val_str = str(val).strip().lower()
            if "scholar" in val_str or "roll" in val_str or "admission" in val_str:
                headers["scholar_number"] = col_idx
            elif "name" in val_str:
                headers["name"] = col_idx
            elif "email" in val_str or "e-mail" in val_str:
                headers["email"] = col_idx
            elif "contact" in val_str or "phone" in val_str or "mobile" in val_str or "number" in val_str:
                if "scholar" not in val_str:
                    headers["contact_number"] = col_idx

        scholar_col = headers.get("scholar_number", 0)
        name_col = headers.get("name", 1)
        email_col = headers.get("email", 2)
        contact_col = headers.get("contact_number", 3)

        for row_idx in range(1, len(rows)):
            row = rows[row_idx]
            if len(row) <= max(scholar_col, name_col):
                continue
                
            scholar_val = row[scholar_col]
            name_val = row[name_col]
            email_val = row[email_col] if len(row) > email_col else ""
            contact_val = row[contact_col] if len(row) > contact_col else ""

            if not scholar_val.strip() or not name_val.strip():
                continue

            students.append({
                "scholar_number": scholar_val.strip(),
                "name": name_val.strip(),
                "email": email_val.strip(),
                "contact_number": contact_val.strip(),
                "image_bytes": None  # CSV files cannot store images
            })

    return students
